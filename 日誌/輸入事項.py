#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工作日誌 — 輸入事項工具 v1.0
============================
交互式 CLI 工具，用於錄入每日工地事項、
管理照片歸屬、生成 Markdown 日誌。

基於 v5.6 模板結構：
  6 欄：時間 | 事項 | 完成度 | 地點 | 工種 | 照片
  地點（必需）+ 工種（可選）獨立兩欄
  固定開工 09:00，收工 17:00（可覆寫）
  支援任務拆分（Task-Splitting Rule v5.6）

保存位置：D:\\AI chat\\openclub\\日誌\\鸿安工程\\
"""

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from typing import Optional

# ── 路徑常數 ──────────────────────────────────────────────
# Windows 路徑（用於 Windows 程式）
WIN_LOG_DIR    = r"D:\AI chat\openclub\日誌\鸿安工程"
WIN_TEMPLATE   = r"D:\AI chat\openclub\日誌\工作日誌模板.md"

# WSL 路徑
WSL_LOG_DIR    = "/mnt/d/AI chat/openclub/日誌/鸿安工程"

# ── 資料模型 ──────────────────────────────────────────────

@dataclass
class Task:
    """一個工作事項"""
    time_range: str       # e.g. "09:00-09:04"
    description: str      # e.g. "開工" or "拆卸舊閥門"
    status: str           # "已完成" | "進行中" | "待辦" | "已取消"
    location: str         # e.g. "星際酒店", "—"
    work_type: str        # e.g. "維修", "—"
    photo_status: str     # "有拍照" | "無照片"
    photos: list[str] = field(default_factory=list)  # 照片檔案路徑清單（絕對路徑）
    item_category: str = ""  # 項位分類（工程對象，如水塔、鍋爐）

    def to_markdown_row(self) -> str:
        """產出 Markdown 表格的一行"""
        loc = self.location if self.location else "—"
        wt = self.work_type if self.work_type else "—"
        ic = self.item_category if self.item_category else "—"
        return f"| <strong>{self.time_range}</strong> | {self.description} | {self.status} | {loc} | {ic} | {wt} | {self.photo_status} |"


@dataclass
class DailyLog:
    """完整的一日工作記錄"""
    date_str: str            # "2026/06/12"
    weekday: str             # "週五"
    person: str              # "千軍破蒼冥"
    tasks: list[Task] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)         # 問題
    risks: list[str] = field(default_factory=list)         # 風險
    followups: list[str] = field(default_factory=list)     # 待跟進
    summary: str = ""                                      # 小復盤
    top3: list[str] = field(default_factory=lambda: ["", "", ""])  # 明日優先 TOP3

    def filename_md(self) -> str:
        d = self.date_str.replace("/", "")
        return f"工作日誌_{d}.md"

    def to_markdown(self) -> str:
        """生成完整的 Markdown 文件內容"""
        lines = []
        lines.append(f"# 工作日誌 {self.date_str}（{self.weekday}）")
        lines.append("")
        lines.append(f"人員：{self.person}")
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("## 時間軸")
        lines.append("")
        lines.append("| 時間 | 事項 | 完成度 | 地點 | 項位分類 | 工種 | 照片 |")
        lines.append("|------|------|--------|------|----------|------|------|")
        for t in self.tasks:
            lines.append(t.to_markdown_row())
        lines.append("")
        lines.append("---")
        lines.append("")

        # 問題/風險/待跟進
        lines.append("## 問題／風險／待跟進")
        lines.append("")
        lines.append("| 問題 | 風險 | 待跟進 |")
        lines.append("|------|------|--------|")
        max_len = max(len(self.notes), len(self.risks), len(self.followups), 1)
        for i in range(max_len):
            n = self.notes[i] if i < len(self.notes) else ""
            r = self.risks[i] if i < len(self.risks) else ""
            f = self.followups[i] if i < len(self.followups) else ""
            lines.append(f"| {n} | {r} | {f} |")
        lines.append("")
        lines.append("---")
        lines.append("")

        # 小復盤
        lines.append("## 小復盤")
        lines.append("")
        if self.summary:
            lines.append(self.summary)
        else:
            lines.append("（無）")
        lines.append("")
        lines.append("---")
        lines.append("")

        # 明日優先 TOP3
        lines.append("## 明日優先 TOP3")
        lines.append("")
        for i, item in enumerate(self.top3, 1):
            lines.append(f"{i}. {item if item else '—'}")
        lines.append("")

        return "\n".join(lines)


# ── 常用選項清單 ──────────────────────────────────────────

COMMON_LOCATIONS = [
    "君悅酒店", "星際酒店", "雅辰酒店",
    "辨公室(文創園)", "辨公室(珠海永發大廈)",
    "南光倉庫", "其他（請輸入）"
]

COMMON_WORK_TYPES = [
    "維修", "落貨", "助手", "安裝", "檢查",
    "搬運", "焊接", "除銹", "其他（請輸入）"
]

COMPLETION_STATUSES = ["已完成", "進行中", "待辦", "已取消"]


# ── 輔助函數 ──────────────────────────────────────────────

def clear_screen():
    """清屏"""
    os.system("cls" if os.name == "nt" else "clear")


def select_from_list(prompt: str, options: list[str], allow_custom: bool = True) -> str:
    """顯示選單讓用戶選擇，返回選中項"""
    print(f"\n{prompt}")
    print("-" * 40)
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")
    print("-" * 40)

    while True:
        try:
            choice = input("請輸入編號 [1-{}]：".format(len(options))).strip()
            if not choice:
                print("  → 輸入不能為空，請重新輸入")
                continue
            idx = int(choice)
            if 1 <= idx <= len(options):
                return options[idx - 1]
            else:
                print(f"  → 請輸入 1~{len(options)} 之間的數字")
        except ValueError:
            if allow_custom:
                print(f"  → 將「{choice}」視為自定義輸入")
                return choice
            print("  → 請輸入有效數字")


def input_non_empty(prompt: str, default: str = "") -> str:
    """讀取非空輸入，可選預設值"""
    while True:
        val = input(prompt).strip()
        if val:
            return val
        if default:
            return default
        print("  → 此欄位不能為空")


def input_optional(prompt: str, default: str = "—") -> str:
    """讀取可選輸入，留空用預設值"""
    val = input(prompt).strip()
    return val if val else default


def get_weekday(date_str: str) -> str:
    """從 YYYY/MM/DD 字串取得星期"""
    parts = date_str.split("/")
    d = date(int(parts[0]), int(parts[1]), int(parts[2]))
    weekdays = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"]
    return weekdays[d.weekday()]


def ensure_dir(path: str):
    """確保目錄存在"""
    os.makedirs(path, exist_ok=True)


# ── 核心互動邏輯 ──────────────────────────────────────────

class DailyLogApp:
    """互動式輸入事項主程式"""

    def __init__(self):
        self.log: Optional[DailyLog] = None

    def run(self):
        """主入口"""
        while True:
            clear_screen()
            print("=" * 50)
            print("   工作日誌 — 輸入事項工具 v1.0")
            print("=" * 50)

            if self.log is None:
                print("\n  [ 尚未開始今日記錄 ]\n")
                print("  1. 開始新工作日誌（開工）")
                print("  2. 查看模板格式")
                print("  0. 退出")
            else:
                task_count = len([t for t in self.log.tasks if t.description != "開工" and t.description != "收工"])
                print(f"\n  日期：{self.log.date_str}（{self.log.weekday}）")
                print(f"  事項數：{task_count} 項（共 {len(self.log.tasks)} 行）")
                print()
                print("  1. 新增事項   2. 檢視/編輯   3. 拆分事項")
                print("  4. 添加照片   5. 填寫備註   6. 收工 & 儲存")
                print("  0. 返回/退出")

            print()
            cmd = input("請選擇操作：").strip()

            if self.log is None:
                if cmd == "1":
                    self._start_new_day()
                elif cmd == "2":
                    self._show_template()
                elif cmd == "0":
                    print("\n  👋 再見！")
                    break
                else:
                    print("\n  → 請先開始新工作日誌（選 1）")
                    input("按 Enter 繼續...")
            else:
                if cmd == "1":
                    self._add_task()
                elif cmd == "2":
                    self._view_edit_tasks()
                elif cmd == "3":
                    self._split_task()
                elif cmd == "4":
                    self._add_photos()
                elif cmd == "5":
                    self._fill_notes()
                elif cmd == "6":
                    self._finish_day()
                elif cmd == "0":
                    self.log = None
                else:
                    input("無效操作，按 Enter 繼續...")

    # ── 子功能 ────────────────────────────────────────────

    def _start_new_day(self):
        """開始新工作日誌"""
        print("\n" + "=" * 50)
        print("   開始新工作日誌（開工）")
        print("=" * 50)

        # 日期
        today = date.today()
        date_str = input(f"日期 [預設 {today.strftime('%Y/%m/%d')}]：").strip()
        if not date_str:
            date_str = today.strftime("%Y/%m/%d")

        weekday = get_weekday(date_str)

        person = input_non_empty("人員 [預設 千軍破蒼冥]：") or "千軍破蒼冥"

        self.log = DailyLog(date_str=date_str, weekday=weekday, person=person)

        # 自動加入開工行 09:00
        self.log.tasks.append(Task(
            time_range="09:00-09:00",
            description="開工",
            status="已完成",
            location="—",
            work_type="—",
            photo_status="無照片",
            item_category=""
        ))

        print(f"\n  ✅ 已建立 {date_str}（{weekday}）工作日誌")
        print(f"  ✅ 已記錄開工 09:00")

        # 詢問是否有第一個事項
        add_now = input("\n是否立即輸入第一個事項？(Y/n)：").strip().lower()
        if add_now != "n":
            self._add_task()

    def _add_task(self):
        """新增事項"""
        if self.log is None:
            return

        print("\n" + "=" * 50)
        print("   新增事項")
        print("=" * 50)

        # 時間範圍
        print("\n--- 時間 ---")
        last_task = self.log.tasks[-1] if self.log.tasks else None
        default_start = last_task.time_range.split("-")[1] if last_task and "-" in last_task.time_range else "09:00"

        start_time = input_non_empty(f"開始時間 [參考上一項結束 {default_start}]：")
        if start_time == "":
            start_time = default_start

        end_time = input_non_empty("結束時間 [HH:MM]：")

        # 事項描述
        print("\n--- 事項描述 ---")
        description = input_non_empty("事項內容：")

        # 完成度
        status = select_from_list("完成度：", COMPLETION_STATUSES, allow_custom=False)

        # 地點
        print("\n--- 地點 ---")
        loc_input = select_from_list("工作地點：", COMMON_LOCATIONS)
        if loc_input == "其他（請輸入）":
            location = input_non_empty("請輸入地點：")
        else:
            location = loc_input

        # 工種
        print("\n--- 工種（可選）---")
        wt_input = select_from_list("工作類型：", COMMON_WORK_TYPES)
        if wt_input == "其他（請輸入）":
            work_type = input("請輸入工種（留空則為 —）：").strip() or "—"
        else:
            work_type = wt_input if wt_input != "—" else "—"

        # 項位分類（工程對象，如水塔、鍋爐；留空則為—）
        print("\n--- 項位分類 ---")
        item_category = input("項位分類（可選，如：水塔、鍋爐）：").strip()

        # 照片
        has_photo = input("\n有拍照嗎？(y/N)：").strip().lower()
        photo_status = "有拍照" if has_photo == "y" else "無照片"

        task = Task(
            time_range=f"{start_time}-{end_time}",
            description=description,
            status=status,
            location=location,
            work_type=work_type,
            photo_status=photo_status,
            item_category=item_category
        )

        self.log.tasks.append(task)
        print(f"\n  ✅ 已新增事項：{description}（{start_time}-{end_time}）")

        # 如果有照片，立即引導添加
        if photo_status == "有拍照":
            add_photo_now = input("\n是否立即添加照片檔案？(y/N)：").strip().lower()
            if add_photo_now == "y":
                self._add_photos_for_task(task)

        input("\n按 Enter 繼續...")

    def _add_photos_for_task(self, task: Task):
        """為特定事項添加照片"""
        while True:
            photo_path = input("照片路徑（留空結束）：").strip()
            if not photo_path:
                break
            if os.path.isfile(photo_path):
                task.photos.append(photo_path)
                print(f"  ✅ 已添加照片：{photo_path}")
            else:
                print(f"  ⚠ 找不到檔案：{photo_path}")

    def _view_edit_tasks(self):
        """檢視及編輯事項清單"""
        if self.log is None or len(self.log.tasks) == 0:
            input("尚無事項，按 Enter 繼續...")
            return

        print("\n" + "=" * 50)
        print("   目前事項清單")
        print("=" * 50)
        print()

        for i, t in enumerate(self.log.tasks):
            print(f"  [{i}] {t.time_range} | {t.description} | {t.status} | {t.category or '—'} | {t.item_category or '—'} | {t.location} | {t.work_type} | {t.photos}")

        print()
        print("  [e] 編輯某項  [d] 刪除某項  [Enter] 返回")
        cmd = input("請選擇：").strip().lower()

        if cmd == "e":
            try:
                idx = int(input("請輸入要編輯的編號：").strip())
                if 0 <= idx < len(self.log.tasks):
                    self._edit_task(idx)
                else:
                    print("  ⚠ 無效編號")
            except ValueError:
                print("  ⚠ 請輸入數字")
        elif cmd == "d":
            try:
                idx = int(input("請輸入要刪除的編號：").strip())
                if 0 <= idx < len(self.log.tasks):
                    deleted = self.log.tasks.pop(idx)
                    print(f"  ✅ 已刪除：{deleted.description}")
                else:
                    print("  ⚠ 無效編號")
            except ValueError:
                print("  ⚠ 請輸入數字")

        input("按 Enter 繼續...")

    def _edit_task(self, idx: int):
        """編輯指定事項"""
        t = self.log.tasks[idx]
        print(f"\n編輯事項 [{idx}]：{t.description}")
        print("（直接按 Enter 保留原值）")

        new_time = input(f"時間 [{t.time_range}]：").strip()
        if new_time:
            t.time_range = new_time

        new_desc = input(f"事項 [{t.description}]：").strip()
        if new_desc:
            t.description = new_desc

        new_status = input(f"完成度 [{t.status}]：").strip()
        if new_status in COMPLETION_STATUSES:
            t.status = new_status

        new_loc = input(f"地點 [{t.location}]：").strip()
        if new_loc:
            t.location = new_loc

        new_wt = input(f"工種 [{t.work_type}]：").strip()
        if new_wt:
            t.work_type = new_wt

        new_ic = input(f"項位分類 [{t.item_category or '—'}]：").strip()
        if new_ic:
            t.item_category = new_ic

        print(f"  ✅ 已更新事項 [{idx}]")

    def _split_task(self):
        """拆分事項（Task-Splitting Rule v5.6）"""
        if self.log is None:
            return

        # 找出可拆分事項（排除開工/收工）
        splittable = [(i, t) for i, t in enumerate(self.log.tasks)
                       if t.description not in ("開工", "收工")]

        if not splittable:
            input("無事項可拆分，按 Enter 繼續...")
            return

        print("\n" + "=" * 50)
        print("   拆分事項")
        print("=" * 50)

        for i, (orig_idx, t) in enumerate(splittable):
            print(f"  [{i}] {t.time_range} | {t.description}")

        try:
            choice = int(input("\n請選擇要拆分的事項編號：").strip())
            if not (0 <= choice < len(splittable)):
                print("  ⚠ 無效編號")
                input("按 Enter 繼續...")
                return
        except ValueError:
            print("  ⚠ 請輸入數字")
            input("按 Enter 繼續...")
            return

        orig_idx, orig_task = splittable[choice]

        print(f"\n拆分：{orig_task.description}")
        print(f"原時間：{orig_task.time_range}")
        print(f"原照片：{len(orig_task.photos)} 張")

        # 詢問拆分點時間
        boundary = input("\n拆分邊界時間 [HH:MM]（如 10:36）：").strip()
        if not boundary:
            print("  ⚠ 必須提供拆分時間")
            input("按 Enter 繼續...")
            return

        # 解析原時間範圍
        try:
            start, end = orig_task.time_range.split("-")
        except ValueError:
            print("  ⚠ 時間格式異常，無法拆分")
            input("按 Enter 繼續...")
            return

        # 第一個子事項名稱
        name1 = input_non_empty("子事項 1 名稱 [原事項前半]：") or orig_task.description
        # 第二個子事項名稱
        name2 = input_non_empty("子事項 2 名稱 [原事項後半]：") or orig_task.description + "（續）"

        # 建立兩個子事項
        task1 = Task(
            time_range=f"{start}-{boundary}",
            description=name1,
            status=orig_task.status,
            location=orig_task.location,
            work_type=orig_task.work_type,
            photo_status=orig_task.photo_status,
            item_category=orig_task.item_category
        )
        task2 = Task(
            time_range=f"{boundary}-{end}",
            description=name2,
            status=orig_task.status,
            location=orig_task.location,
            work_type=orig_task.work_type,
            photo_status=orig_task.photo_status,
            item_category=orig_task.item_category
        )

        # 照片重新分配
        if orig_task.photos:
            print(f"\n原事項有 {len(orig_task.photos)} 張照片需要重新分配：")
            for pi, pp in enumerate(orig_task.photos):
                print(f"  相{pi+1}：{os.path.basename(pp)}")
                assign = input(f"  歸屬？(1={name1}, 2={name2}) [預設 1]：").strip()
                if assign == "2":
                    task2.photos.append(pp)
                else:
                    task1.photos.append(pp)
            # 更新照片狀態
            task1.photo_status = "有拍照" if task1.photos else "無照片"
            task2.photo_status = "有拍照" if task2.photos else "無照片"

        # 替換
        self.log.tasks[orig_idx] = task1
        self.log.tasks.insert(orig_idx + 1, task2)

        print(f"\n  ✅ 已拆分：{orig_task.description}")
        print(f"     → {task1.time_range} | {task1.description}")
        print(f"     → {task2.time_range} | {task2.description}")

        input("按 Enter 繼續...")

    def _add_photos(self):
        """為已存在的事項添加照片"""
        if self.log is None:
            return

        # 找出有拍照的事項
        photo_tasks = [(i, t) for i, t in enumerate(self.log.tasks)
                        if t.description not in ("開工", "收工")]

        if not photo_tasks:
            input("尚無可添加照片的事項，按 Enter 繼續...")
            return

        print("\n" + "=" * 50)
        print("   添加照片")
        print("=" * 50)

        for i, (idx, t) in enumerate(photo_tasks):
            pcount = len(t.photos)
            print(f"  [{i}] {t.time_range} | {t.description} （已有 {pcount} 張照片）")

        try:
            choice = int(input("\n選擇要歸屬照片的事項編號：").strip())
            if not (0 <= choice < len(photo_tasks)):
                print("  ⚠ 無效編號")
                input("按 Enter 繼續...")
                return
        except ValueError:
            print("  ⚠ 請輸入數字")
            input("按 Enter 繼續...")
            return

        _, task = photo_tasks[choice]
        self._add_photos_for_task(task)
        task.photo_status = "有拍照" if task.photos else task.photo_status
        input("按 Enter 繼續...")

    def _fill_notes(self):
        """填寫底部五個區塊"""
        if self.log is None:
            return

        print("\n" + "=" * 50)
        print("   填寫底部區塊")
        print("=" * 50)

        print("\n--- 問題（逐項輸入，留空結束）---")
        while True:
            item = input(f"  問題 #{len(self.log.notes) + 1}：").strip()
            if not item:
                break
            self.log.notes.append(item)

        print("\n--- 風險（逐項輸入，留空結束）---")
        while True:
            item = input(f"  風險 #{len(self.log.risks) + 1}：").strip()
            if not item:
                break
            self.log.risks.append(item)

        print("\n--- 待跟進（逐項輸入，留空結束）---")
        while True:
            item = input(f"  待跟進 #{len(self.log.followups) + 1}：").strip()
            if not item:
                break
            self.log.followups.append(item)

        print("\n--- 小復盤 ---")
        summary = input("（多行輸入，輸入 END 結束）：\n").strip()
        lines = []
        while True:
            line = input()
            if line.strip().upper() == "END":
                break
            lines.append(line)
        self.log.summary = "\n".join(lines) if lines else summary

        print("\n--- 明日優先 TOP3 ---")
        for i in range(3):
            item = input(f"  TOP{i+1}：").strip()
            if item:
                self.log.top3[i] = item

        print("\n  ✅ 備註已更新")
        input("按 Enter 繼續...")

    def _finish_day(self):
        """收工 & 儲存"""
        if self.log is None:
            return

        # 檢查是否已有收工行
        has_end = any(t.description == "收工" for t in self.log.tasks)

        if not has_end:
            print("\n" + "=" * 50)
            print("   收工")
            print("=" * 50)

            last_end = self.log.tasks[-1].time_range.split("-")[1] if self.log.tasks else "17:00"
            end_time = input(f"收工時間 [預設 17:00，上一次事項結束 {last_end}]：").strip()
            if not end_time:
                end_time = "17:00"

            self.log.tasks.append(Task(
                time_range=f"{end_time}-{end_time}",
                description="收工",
                status="已完成",
                location="—",
                work_type="—",
                photo_status="無照片",
                item_category=""
            ))
            print(f"  ✅ 已記錄收工 {end_time}")

        # 預覽
        print("\n" + "=" * 50)
        print("   日誌預覽")
        print("=" * 50)
        print()
        print(self.log.to_markdown())
        print()

        confirm = input("是否儲存？(Y/n)：").strip().lower()
        if confirm == "n":
            print("  ⚠ 未儲存")
            input("按 Enter 繼續...")
            return

        # 儲存 Markdown
        ensure_dir(WSL_LOG_DIR)
        md_path = os.path.join(WSL_LOG_DIR, self.log.filename_md())
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(self.log.to_markdown())
        print(f"\n  ✅ 已儲存：{md_path}")

        # 重置
        self.log = None
        print("\n  ✅ 工作日誌已完成！")
        input("按 Enter 繼續...")

    def _show_template(self):
        """顯示模板格式"""
        print("\n" + "=" * 50)
        print("   工作日誌模板 v5.6 格式")
        print("=" * 50)
        print()
        print("| 時間 | 事項 | 完成度 | 分類 | 項位分類 | 地點 | 工種 | 照片 |")
        print("|------|------|--------|------|----------|------|------|------|")
        print("| 09:00-09:04 | 開工 | 已完成 | 個人 | — | — | — | 無照片 |")
        print("| 09:04-10:11 | 更換零件 | 已完成 | 維護 | 水泵 | 星際酒店 | 維修 | 有拍照 |")
        print("| 17:00-17:00 | 收工 | 已完成 | 個人 | — | — | — | 無照片 |")
        print()
        print("底部區塊：")
        print("  問題／風險／待跟進（表格）")
        print("  小復盤（文字）")
        print("  明日優先 TOP3（列表）")
        print()
        print("預設存檔：D:\\AI chat\\openclub\\日誌\\鸿安工程\\")
        input("按 Enter 繼續...")


# ── Markdown → XLSX 匯出功能 ────────────────────────────


def parse_markdown_log(md_path: str) -> dict:
    """解析工作日誌 Markdown 文件，返回結構化資料"""
    with open(md_path, "r", encoding="utf-8") as f:
        text = f.read()

    data = {
        "date_str": "",
        "weekday": "",
        "person": "",
        "tasks": [],
        "notes": [],
        "risks": [],
        "followups": [],
        "summary": "",
        "top3": [],
        "attendance": []
    }

    # 標題行：工作日誌 YYYY/MM/DD（週X）（可能有 # 前綴，也可能沒有）
    m = re.search(r"#?\s*工作日誌\s*(\d{4}/\d{2}/\d{2})（(.+?)）", text)
    if m:
        data["date_str"] = m.group(1)
        data["weekday"] = m.group(2)

    # 人員
    m = re.search(r"人員：(.+)", text)
    if m:
        data["person"] = m.group(1).strip()

    # ── 時間軸表格 ──
    # 找到 ## 時間軸 到下一 ## 之間的表格行
    timeline_section = re.search(
        r"## 時間軸.*?\n((?:\|.+\|\s*\n)+)", text, re.DOTALL
    )
    if timeline_section:
        table_text = timeline_section.group(1)
        rows = table_text.strip().split("\n")
        # 跳過表頭（第1行）和分隔線（第2行）
        for row in rows[2:]:
            row = row.strip()
            if not row.startswith("|"):
                continue
            cells = [c.strip() for c in row.split("|")[1:-1]]
            if len(cells) < 6:
                continue
            # 移除 HTML 標籤 <strong>...</strong>
            time_range = re.sub(r"</?strong>", "", cells[0]).strip()
            # 跳過分隔行（時間為 —）
            if time_range in ("—", "", "-"):
                continue
            # Auto-detect: 8-col (含項位分類) vs 7-col (含分類) vs 6-col
            # 8-col: 時間|事項|完成度|分類|項位分類|地點|工種|照片
            # 7-col: 時間|事項|完成度|分類|地點|工種|照片
            # 6-col: 時間|事項|完成度|地點|工種|照片
            if len(cells) >= 8:
                data["tasks"].append({
                    "time_range": time_range,
                    "description": cells[1],
                    "status": cells[2],
                    "category": cells[3],
                    "item_category": cells[4],
                    "project": "",
                    "personnel": "",
                    "location": cells[5],
                    "work_type": cells[6],
                    "photo_status": cells[7]
                })
            elif len(cells) >= 7:
                data["tasks"].append({
                    "time_range": time_range,
                    "description": cells[1],
                    "status": cells[2],
                    "category": cells[3],
                    "item_category": "",
                    "project": "",
                    "personnel": "",
                    "location": cells[4],
                    "work_type": cells[5],
                    "photo_status": cells[6]
                })
            else:
                data["tasks"].append({
                    "time_range": time_range,
                    "description": cells[1],
                    "status": cells[2],
                    "category": "",
                    "item_category": "",
                    "project": "",
                    "personnel": "",
                    "location": cells[3],
                    "work_type": cells[4],
                    "photo_status": cells[5]
                })

    # ── 問題／風險／待跟進 ──
    issues_section = re.search(
        r"## 問題／風險／待跟進.*?\n((?:\|.+\|\s*\n)+)", text, re.DOTALL
    )
    if issues_section:
        table_text = issues_section.group(1)
        rows = table_text.strip().split("\n")
        for row in rows[2:]:  # 跳過表頭和分隔線
            row = row.strip()
            if not row.startswith("|"):
                continue
            cells = [c.strip() for c in row.split("|")[1:-1]]
            if len(cells) >= 3:
                if cells[0]:
                    data["notes"].append(cells[0])
                if cells[1]:
                    data["risks"].append(cells[1])
                if cells[2]:
                    data["followups"].append(cells[2])

    # ── 小復盤 ──
    m = re.search(r"## 小復盤\s*\n(.+?)(?:\n---|\n##)", text, re.DOTALL)
    if m:
        summary = m.group(1).strip()
        if summary and summary != "（無）":
            data["summary"] = summary

    # ── 明日優先 TOP3 ──
    m = re.search(r"## 明日優先 TOP3\s*\n(.+?)$", text, re.DOTALL)
    if m:
        top3_text = m.group(1).strip()
        for line in top3_text.split("\n"):
            line = line.strip()
            # 匹配 "1. xxx" 或 "1. xxx"
            t = re.sub(r"^\d+\.\s*", "", line).strip()
            if t and t != "—":
                data["top3"].append(t)

    return data


def generate_xlsx(data: dict, output_path: str):
    """根據結構化資料產生 XLSX 工作日誌"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 通用樣式 ──
    header_font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    section_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    label_font = Font(name="Microsoft YaHei", bold=True, size=10)
    body_font = Font(name="Microsoft YaHei", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Auto-detect: does data have category? and item_category?
    has_cat = any(t.get("category", "") for t in data.get("tasks", []))
    has_ic = any(t.get("item_category", "") for t in data.get("tasks", []))
    has_project = any(t.get("project", "") for t in data.get("tasks", []))
    has_personnel = any(t.get("personnel", "") for t in data.get("tasks", []))

    # 按網站規範順序組裝輸出欄位（有先加、冇就唔出）
    selected_fields = [
        ("時間", "time_range", 10),
        ("事項", "description", 22),
        ("完成度", "status", 7),
    ]
    if has_cat:
        selected_fields.append(("分類", "category", 7))
    if has_ic:
        selected_fields.append(("項位分類", "item_category", 10))
    if has_ic and (has_project or has_personnel):
        if has_project:
            selected_fields.append(("項目", "project", 14))
        if has_personnel:
            selected_fields.append(("人員", "personnel", 8))
    selected_fields += [
        ("地點", "location", 10),
        ("工種", "work_type", 7),
        ("照片", "photo_status", 8),
    ]

    col_headers = [f[0] for f in selected_fields]
    col_widths = [f[2] for f in selected_fields]
    field_keys = [f[1] for f in selected_fields]
    ncol = len(col_headers)

    # ═══════════════════════════════════════════════
    # Sheet 1: 時間軸
    # ═══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "時間軸"

    row = 1

    # 標題（跨 N 欄置中）
    title = f"工作日誌 {data['date_str']}（{data['weekday']}）"
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    cell = ws1.cell(row=row, column=1, value=title)
    cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # 欄標題 + 欄寬
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws1.cell(row=row, column=col_idx, value=h)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(name="Microsoft YaHei", bold=True, size=9, color="FFFFFF")
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws1.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    # 凍結標題列（捲動時表頭固定）
    ws1.freeze_panes = f"A{row + 1}"
    # 啟用自動篩選
    col_letter = get_column_letter(ncol)
    ws1.auto_filter.ref = f"A{row}:{col_letter}{row + len(data.get('tasks', []))}"
    header_row = row
    row += 1

    # 任務行
    for t in data.get("tasks", []):
        vals = []
        for fk in field_keys:
            if fk == "time_range":
                v = t.get("time_range", "")
            elif fk == "description":
                v = t.get("description", "")
            elif fk == "status":
                v = t.get("status", "")
            elif fk == "category":
                v = t.get("category", "—") or "—"
            elif fk == "item_category":
                v = t.get("item_category", "—") or "—"
            elif fk == "project":
                v = t.get("project", "—") or "—"
            elif fk == "personnel":
                v = t.get("personnel", "—") or "—"
            elif fk == "location":
                v = t.get("location", "—") or "—"
            elif fk == "work_type":
                v = t.get("work_type", "—") or "—"
            elif fk == "photo_status":
                v = t.get("photo_status", "無照片") or "無照片"
            else:
                v = ""
            vals.append(v)
        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col_idx, value=val)
            cell.font = Font(name="Microsoft YaHei", size=9)
            cell.alignment = wrap_align
            cell.border = thin_border
            # 交替行底色
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        row += 1

    # ═══════════════════════════════════════════════
    # Sheet 2: 備註
    # ═══════════════════════════════════════════════
    ws2 = wb.create_sheet(title="備註")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 42

    row = 1

    # 標題
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws2.cell(row=row, column=1, value=title)
    cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    # ── 問題 ──
    if data.get("notes"):
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws2.cell(row=row, column=1, value="⚠️ 問題")
        cell.font = section_font
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        row += 1
        for note in data["notes"]:
            ws2.cell(row=row, column=1, value="問題").font = label_font
            ws2.cell(row=row, column=1).border = thin_border
            ws2.cell(row=row, column=2, value=note).font = body_font
            ws2.cell(row=row, column=2).alignment = wrap_align
            ws2.cell(row=row, column=2).border = thin_border
            row += 1
        row += 1

    # ── 風險 ──
    if data.get("risks"):
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws2.cell(row=row, column=1, value="⚠️ 風險")
        cell.font = section_font
        cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        row += 1
        for risk in data["risks"]:
            ws2.cell(row=row, column=1, value="風險").font = label_font
            ws2.cell(row=row, column=1).border = thin_border
            ws2.cell(row=row, column=2, value=risk).font = body_font
            ws2.cell(row=row, column=2).alignment = wrap_align
            ws2.cell(row=row, column=2).border = thin_border
            row += 1
        row += 1

    # ── 待跟進 ──
    if data.get("followups"):
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws2.cell(row=row, column=1, value="🔁 待跟進")
        cell.font = section_font
        cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
        row += 1
        for f in data["followups"]:
            ws2.cell(row=row, column=1, value="待跟進").font = label_font
            ws2.cell(row=row, column=1).border = thin_border
            ws2.cell(row=row, column=2, value=f).font = body_font
            ws2.cell(row=row, column=2).alignment = wrap_align
            ws2.cell(row=row, column=2).border = thin_border
            row += 1
        row += 1

    # ── 小復盤 ──
    if data.get("summary"):
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws2.cell(row=row, column=1, value="💡 小復盤")
        cell.font = section_font
        cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        row += 1
        ws2.cell(row=row, column=1, value="復盤").font = label_font
        ws2.cell(row=row, column=1).border = thin_border
        ws2.cell(row=row, column=2, value=data["summary"]).font = body_font
        ws2.cell(row=row, column=2).alignment = wrap_align
        ws2.cell(row=row, column=2).border = thin_border
        row += 2

    # ── 明日優先 TOP3 ──
    if data.get("top3"):
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws2.cell(row=row, column=1, value="📋 明日優先 TOP3")
        cell.font = section_font
        cell.fill = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
        row += 1
        for i, item in enumerate(data["top3"], 1):
            ws2.cell(row=row, column=1, value=f"TOP{i}").font = label_font
            ws2.cell(row=row, column=1).border = thin_border
            ws2.cell(row=row, column=2, value=item).font = body_font
            ws2.cell(row=row, column=2).alignment = wrap_align
            ws2.cell(row=row, column=2).border = thin_border
            row += 1

    # 儲存
    wb.save(output_path)
    return output_path


# ── 欄位名 → 資料鍵 對照（對應網站 OUT_FIELDS）────────────────────
# 網站（輸入事項.html v1.5+）輸出模板可自訂欄位順序／子集：
#   時間 | 事項 | 完成度 | 分類 | 項位分類 | 項目 | 人員 | 地點 | 工種 | 照片
# 讀取時以「欄位名」為準按欄填充，避免固定欄位位置錯位。
XLSX_HEADER_MAP = {
    "時間": "time_range",
    "事項": "description",
    "完成度": "status",
    "分類": "category",
    "項位分類": "item_category",
    "項目": "project",
    "人員": "personnel",
    "地點": "location",
    "工種": "work_type",
    "照片": "photo_status",
}

# 時間軸欄位規範順序（網站 OUT_FIELDS 順序）：(表頭, 資料鍵, 欄寬)
XLSX_COL_ORDER = [
    ("時間", "time_range", 10),
    ("事項", "description", 22),
    ("完成度", "status", 7),
    ("分類", "category", 7),
    ("項位分類", "item_category", 10),
    ("項目", "project", 14),
    ("人員", "personnel", 8),
    ("地點", "location", 10),
    ("工種", "work_type", 7),
    ("照片", "photo_status", 8),
]

# 各欄位預設值（欄位唔喺輸出模板入面時使用）
TASK_FIELD_DEFAULTS = {
    "time_range": "",
    "description": "",
    "status": "已完成",
    "category": "",
    "item_category": "",
    "project": "",
    "personnel": "",
    "location": "—",
    "work_type": "—",
    "photo_status": "無照片",
}


def _norm_header(h) -> str:
    """正規化表頭文字：去空白／全形空格／NBSP"""
    if h is None:
        return ""
    return str(h).strip().replace("\u3000", "").replace("\xa0", "")


def detect_header_row(ws, max_scan: int = 8):
    """
    掃描工作表首幾行，搵出表頭行並建立「資料鍵 → 欄號」對照。

    Returns:
        (header_row_index, col_map)
        col_map: {field_key: column_number}；完全搵唔到表頭時為空 dict。
    """
    best_row = None
    best_map: dict = {}
    best_score = 0
    for r in range(1, min(max_scan, ws.max_row) + 1):
        col_map: dict = {}
        for c in range(1, ws.max_column + 1):
            h = _norm_header(ws.cell(row=r, column=c).value)
            if not h:
                continue
            key = XLSX_HEADER_MAP.get(h)
            if key is None:
                # 子字串匹配（例如「項位分類（工程對象）」）
                for label, k in XLSX_HEADER_MAP.items():
                    if label and label in h:
                        key = k
                        break
            if key and key not in col_map:
                col_map[key] = c
        score = len(col_map)
        if score > best_score:
            best_score = score
            best_row = r
            best_map = col_map
    return best_row, best_map


# ════════════════════════════════════════════════════════════════
# 功能二：XLSX → PDF（weasyprint）
# ════════════════════════════════════════════════════════════════


def read_xlsx_data(xlsx_path: str) -> dict:
    """Read a daily log XLSX and return the same data dict structure."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)

    data = {
        "date_str": "",
        "weekday": "",
        "person": "",
        "tasks": [],
        "notes": [],
        "risks": [],
        "followups": [],
        "summary": "",
        "top3": [],
        "attendance": []
    }

    # Sheet 1: 時間軸
    ws1 = wb["時間軸"]

    # 第一行是標題：工作日誌 YYYY/MM/DD（週X）（網站可能用 - 分隔日期、冇週X）
    title_cell = ws1.cell(row=1, column=1).value or ""
    m = re.search(r"(\d{4}[/-]\d{2}[/-]\d{2})(?:（(.+?)）)?", str(title_cell))
    if m:
        data["date_str"] = m.group(1).replace("-", "/")
        data["weekday"] = m.group(2) or ""

    # ── 讀取表頭欄位名，按欄填充內容 ──
    # 網站（輸入事項.html v1.5+）輸出模板可自訂欄位順序／子集，所以
    # 唔可以再用固定欄位位置，必須以「欄位名」定位每一欄再填充。
    header_row_idx, col_map = detect_header_row(ws1)

    if len(col_map) >= 2:
        # ── 新格式：按欄位名填充 ──
        for row in range(header_row_idx + 1, ws1.max_row + 1):
            time_col = col_map.get("time_range", 1)
            time_val = ws1.cell(row=row, column=time_col).value
            if time_val is None or str(time_val).strip() in ("", "—"):
                continue
            task = {}
            for _label, key in XLSX_HEADER_MAP.items():
                col = col_map.get(key)
                if col:
                    val = ws1.cell(row=row, column=col).value
                    task[key] = str(val).strip() if val is not None else ""
                else:
                    task[key] = TASK_FIELD_DEFAULTS[key]
            if not task["status"]:
                task["status"] = "已完成"
            data["tasks"].append(task)

        # 人員：優先從「人員」欄第一行非空值取值，否則預設「傑」
        person = "傑"
        if "personnel" in col_map:
            pcol = col_map["personnel"]
            for row in range(header_row_idx + 1, ws1.max_row + 1):
                v = ws1.cell(row=row, column=pcol).value
                v = str(v).strip() if v is not None else ""
                if v and v not in ("—", ""):
                    person = v
                    break
        data["person"] = person
    else:
        # ── 舊格式 fallback：固定欄位位置（8/7/6 欄）──
        # 8-col: 時間|事項|完成度|分類|項位分類|地點|工種|照片
        # 7-col: 時間|事項|完成度|分類|地點|工種|照片
        # 6-col: 時間|事項|完成度|地點|工種|照片
        header_row = ws1.cell(row=3, column=4).value or ""
        header_5 = str(ws1.cell(row=3, column=5).value or "")
        is_7col = "分類" in str(header_row)
        is_8col = is_7col and ("項位" in header_5 or "分類" in header_5)

        for row in range(4, ws1.max_row + 1):
            time_val = ws1.cell(row=row, column=1).value
            if not time_val or str(time_val).strip() in ("", "—"):
                continue
            if is_8col:
                data["tasks"].append({
                    "time_range": str(ws1.cell(row=row, column=1).value or "").strip(),
                    "description": str(ws1.cell(row=row, column=2).value or "").strip(),
                    "status": str(ws1.cell(row=row, column=3).value or "").strip(),
                    "category": str(ws1.cell(row=row, column=4).value or "維護").strip(),
                    "item_category": str(ws1.cell(row=row, column=5).value or "").strip(),
                    "project": "",
                    "personnel": "",
                    "location": str(ws1.cell(row=row, column=6).value or "—").strip(),
                    "work_type": str(ws1.cell(row=row, column=7).value or "—").strip(),
                    "photo_status": str(ws1.cell(row=row, column=8).value or "無照片").strip()
                })
            elif is_7col:
                data["tasks"].append({
                    "time_range": str(ws1.cell(row=row, column=1).value or "").strip(),
                    "description": str(ws1.cell(row=row, column=2).value or "").strip(),
                    "status": str(ws1.cell(row=row, column=3).value or "").strip(),
                    "category": str(ws1.cell(row=row, column=4).value or "維護").strip(),
                    "item_category": "",
                    "project": "",
                    "personnel": "",
                    "location": str(ws1.cell(row=row, column=5).value or "—").strip(),
                    "work_type": str(ws1.cell(row=row, column=6).value or "—").strip(),
                    "photo_status": str(ws1.cell(row=row, column=7).value or "無照片").strip()
                })
            else:
                data["tasks"].append({
                    "time_range": str(ws1.cell(row=row, column=1).value or "").strip(),
                    "description": str(ws1.cell(row=row, column=2).value or "").strip(),
                    "status": str(ws1.cell(row=row, column=3).value or "").strip(),
                    "category": "",
                    "item_category": "",
                    "project": "",
                    "personnel": "",
                    "location": str(ws1.cell(row=row, column=4).value or "—").strip(),
                    "work_type": str(ws1.cell(row=row, column=5).value or "—").strip(),
                    "photo_status": str(ws1.cell(row=row, column=6).value or "無照片").strip()
                })

        # 人員：從 XLSX 無法直接還原，預設為「傑」
        data["person"] = "傑"

    # Sheet 2: 備註
    if "備註" in wb.sheetnames:
        ws2 = wb["備註"]
        section = None
        for row in range(1, ws2.max_row + 1):
            a = str(ws2.cell(row=row, column=1).value or "").strip()
            b = str(ws2.cell(row=row, column=2).value or "").strip()
            # Section headers: merged cells (col B empty), contain emoji prefix
            if not b:
                if "問題" in a:
                    section = "notes"
                elif "風險" in a:
                    section = "risks"
                elif "考勤" in a:
                    section = "attendance"
                    continue
                elif "待跟進" in a:
                    section = "followups"
                elif "小復盤" in a:
                    section = "summary"
                    continue
                elif "TOP" in a:
                    section = "top3"
            # Data rows: col B has actual content
            elif b:
                if section == "notes":
                    data["notes"].append(b)
                elif section == "risks":
                    data["risks"].append(b)
                elif section == "attendance":
                    if a and a != "人員":
                        data["attendance"].append({"name": a, "status": b})
                elif section == "followups":
                    data["followups"].append(b)
                elif section == "summary":
                    data["summary"] = b
                elif section == "top3":
                    data["top3"].append(b)

    wb.close()
    return data


def build_category_summary_html(tasks: list) -> str:
    """Build a 4-column category summary table (Design A).
     各分類：🔧維護 🔍檢查 🤝助手 👤個人
     每個分類欄位直排列出事項+時間。"""
    CATEGORIES = ["維護", "檢查", "助手", "個人"]
    CAT_ICONS = {"維護": "🔧", "檢查": "🔍", "助手": "🤝", "個人": "👤"}
    CAT_COLORS = {"維護": "#27ae60", "檢查": "#2980b9", "助手": "#e67e22", "個人": "#8e44ad"}

    # 分類 tasks 入 dict
    grouped: dict[str, list] = {c: [] for c in CATEGORIES}
    for t in tasks:
        cat = t.get("category", "").strip()
        if cat in grouped:
            grouped[cat].append(t)
        else:
            # 冇分類 → 預設放入個人
            if t.get("description") in ("開工", "收工", "午休"):
                grouped["個人"].append(t)

    # 搵最大行數（最多事項嘅分類有幾多項）
    max_rows = max((len(grouped[c]) for c in CATEGORIES), default=1)

    # 逐行砌 HTML
    rows_html = ""
    for r in range(max_rows):
        cells = ""
        for c in CATEGORIES:
            items = grouped.get(c, [])
            if r < len(items):
                t = items[r]
                cells += f"""\
            <td style="width:25%;vertical-align:top;border:1px solid #ddd;padding:6px;">
                <div style="font-size:10pt;font-weight:bold;">{t['time_range']}</div>
                <div style="font-size:10pt;color:#555;">{t['description']}</div>
            </td>"""
            else:
                cells += """\
            <td style="width:25%;vertical-align:top;border:1px solid #ddd;padding:6px;">
                <div style="font-size:10pt;color:#ccc;">—</div>
            </td>"""
        rows_html += f"        <tr>\n{cells}\n        </tr>\n"

    # Header row with category icons + colors
    head_cells = ""
    for c in CATEGORIES:
        color = CAT_COLORS.get(c, "#333")
        head_cells += f"""\
            <th style="width:25%;background-color:{color};color:#fff;padding:8px 6px;text-align:center;font-weight:bold;font-size:11pt;">
                {CAT_ICONS.get(c, "")} {c}
            </th>"""

    return f"""\
<h2>事項分類摘要</h2>
<table style="width:100%;border-collapse:collapse;margin-bottom:20px;font-size:10pt;">
    <thead>
        <tr>
{head_cells}
        </tr>
    </thead>
    <tbody>
{rows_html}
    </tbody>
</table>"""


def xlsx_data_to_html(data: dict) -> str:
    """Convert parsed XLSX data dict to HTML for PDF generation."""
    title = f"工作日誌 {data['date_str']}（{data['weekday']}）"
    person = data.get("person", "傑")

    # Build task rows — 動態欄位：先偵測有咩欄，再按欄填充
    task_rows = ""
    has_category = False
    has_item_category = False
    has_project = False
    has_personnel = False
    for t in data.get("tasks", []):
        if (t.get("category") or "").strip():
            has_category = True
        if (t.get("item_category") or "").strip():
            has_item_category = True
        if (t.get("project") or "").strip():
            has_project = True
        if (t.get("personnel") or "").strip():
            has_personnel = True

    col_defs = [("時間", "10%"), ("事項", "16%"), ("完成度", "7%")]
    if has_category:
        col_defs.append(("分類", "7%"))
    if has_item_category:
        col_defs.append(("項位分類", "9%"))
    if has_project:
        col_defs.append(("項目", "10%"))
    if has_personnel:
        col_defs.append(("人員", "7%"))
    col_defs += [("地點", "11%"), ("工種", "9%"), ("照片", "24%")]

    def _task_cell(t, header):
        if header == "時間":
            return t['time_range']
        if header == "事項":
            return t['description']
        if header == "完成度":
            return t['status']
        if header == "分類":
            return (t.get("category") or "").strip() or "—"
        if header == "項位分類":
            return (t.get("item_category") or "").strip() or "—"
        if header == "項目":
            return (t.get("project") or "").strip() or "—"
        if header == "人員":
            return (t.get("personnel") or "").strip() or "—"
        if header == "地點":
            return (t.get("location") or "—") or "—"
        if header == "工種":
            return (t.get("work_type") or "—") or "—"
        if header == "照片":
            return t['photo_status']
        return ""

    for t in data.get("tasks", []):
        cells = "".join(f"<td>{_task_cell(t, h)}</td>" for h, _w in col_defs)
        task_rows += f"<tr>\n{cells}\n</tr>\n"

    # Issue table
    issue_rows = ""
    notes_list = data.get("notes", []) or []
    risks_list = data.get("risks", []) or []
    followups_list = data.get("followups", []) or []
    max_issues = max(len(notes_list), len(risks_list), len(followups_list), 1)
    for i in range(max_issues):
        n = notes_list[i] if i < len(notes_list) else ""
        r = risks_list[i] if i < len(risks_list) else ""
        f = followups_list[i] if i < len(followups_list) else ""
        issue_rows += f"""\
            <tr>
                <td>{n}</td>
                <td>{r}</td>
                <td>{f}</td>
            </tr>"""

    # Summary
    summary = data.get("summary", "") or ""
    summary_html = f"<p>{summary.replace(chr(10), '<br>')}</p>" if summary else "<p>（無）</p>"

    # Top3
    top3_list = data.get("top3", []) or []
    top3_html = ""
    for i, item in enumerate(top3_list, 1):
        item = item.strip()
        top3_html += f"<li>{item if item else '—'}</li>\n"
    if not top3_html:
        top3_html = "<li>—</li>\n<li>—</li>\n<li>—</li>\n"

    # 動態地點摘要
    all_locs = set()
    for t in data.get("tasks", []):
        loc = t.get("location", "") or ""
        if loc and loc != "—":
            all_locs.add(loc)
    location_summary = "／".join(sorted(all_locs)) if all_locs else "—"

    # 考勤摘要（網站輸出「👥 考勤」區塊時顯示）
    attendance = data.get("attendance") or []
    attendance_html = ""
    if attendance:
        atts = " ｜ ".join(
            f"{a.get('name', '')}:{a.get('status', '')}"
            for a in attendance if a.get("name")
        )
        if atts:
            attendance_html = f'<p class="meta">考勤：{atts}</p>'

    # Category summary table (Design A)
    category_summary = ""
    if has_category:
        category_summary = build_category_summary_html(data.get("tasks", []))

    # Time-axis table headers — 動態對應 col_defs
    timeline_headers = "".join(
        f'<th style="width:{w}">{h}</th>' for h, w in col_defs
    )
    category_legend = '<p style="font-size:9pt;color:#666;margin-top:-15px;margin-bottom:20px;">分類：🔧維護 ｜ 🔍檢查 ｜ 🤝助手 ｜ 👤個人</p>' if has_category else ""

    return f"""\
<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="utf-8">
<style>
    @page {{ size: A4; margin: 2cm; }}
    body {{ font-family: 'Noto Sans SC','Microsoft YaHei','PingFang SC',sans-serif; font-size: 12pt; color: #333; }}
    h1 {{ text-align: center; font-size: 22pt; color: #2c3e50; border-bottom: 3px solid #e94560; padding-bottom: 10px; margin-bottom: 20px; }}
    h2 {{ font-size: 16pt; color: #2c3e50; border-left: 4px solid #e94560; padding-left: 10px; margin-top: 25px; margin-bottom: 10px; }}
    .meta {{ text-align: center; font-size: 11pt; color: #666; margin-bottom: 15px; }}
    table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 10pt; }}
    th {{ background-color: #2c3e50; color: #fff; padding: 8px 6px; text-align: center; font-weight: bold; }}
    td {{ padding: 6px; border: 1px solid #ddd; }}
    tr:nth-child(even) td {{ background-color: #f8f9fa; }}
    .section-box {{ background: #fafafa; border: 1px solid #e0e0e0; border-radius: 4px; padding: 12px 15px; margin-bottom: 15px; }}
    .note-box {{ background: #f0f7ff; border-left: 4px solid #3498db; padding: 10px 15px; margin-bottom: 15px; }}
    .footer {{ text-align: center; font-size: 9pt; color: #95a5a6; margin-top: 30px; border-top: 1px solid #eee; padding-top: 10px; }}
    ul {{ margin: 0; padding-left: 20px; }}
    li {{ margin-bottom: 4px; }}
</style>
</head>
<body>
<h1>{title}</h1>
<p class="meta">人員：{person} ｜ 地點：{location_summary}</p>
{attendance_html}

{category_summary}

<h2>時間軸</h2>
<table>
    <thead>
        <tr>
{timeline_headers}
        </tr>
    </thead>
    <tbody>
{task_rows}
    </tbody>
</table>
{category_legend}

<h2>問題／風險／待跟進</h2>
<table>
    <thead>
        <tr>
            <th>問題</th>
            <th>風險</th>
            <th>待跟進</th>
        </tr>
    </thead>
    <tbody>
{issue_rows}
    </tbody>
</table>

<h2>小復盤</h2>
<div class="note-box">
{summary_html}
</div>

<h2>明日優先 TOP3</h2>
<div class="section-box">
<ul>
{top3_html}
</ul>
</div>

<div class="footer">由 Hermes Agent · 輸入事項工具 自動產生</div>
</body>
</html>"""


def xlsx_to_pdf(xlsx_path: str, output_path: str = None):
    """Convert a daily log XLSX to PDF using WeasyPrint."""
    import weasyprint

    if output_path is None:
        output_path = os.path.splitext(xlsx_path)[0] + ".pdf"

    print(f"📖 讀取 XLSX：{xlsx_path}")
    data = read_xlsx_data(xlsx_path)
    print(f"   ✓ 日期：{data['date_str']}（{data['weekday']}）")
    print(f"   ✓ 事項：{len(data['tasks'])} 項")

    html = xlsx_data_to_html(data)
    doc = weasyprint.HTML(string=html)
    doc.write_pdf(output_path)
    print(f"✅ 已輸出 PDF：{output_path}")
    return output_path


def merge_xlsx_pdf(xlsx_paths: list[str], output_path: str = None):
    """Merge multiple daily log XLSX files into a single PDF."""
    import weasyprint

    if output_path is None:
        # 預設名稱取自第一個檔案日期範圍
        base = os.path.dirname(xlsx_paths[0]) if len(xlsx_paths) > 1 else os.path.dirname(xlsx_paths[0])
        output_path = os.path.join(base, "工作日誌_合併.pdf")

    print(f"📖 合併 {len(xlsx_paths)} 個 XLSX 檔案 → PDF")
    print("-" * 40)

    all_html_parts = []
    for xlsx_path in sorted(xlsx_paths):
        data = read_xlsx_data(xlsx_path)
        html = xlsx_data_to_html(data)
        # Add page break between files
        if all_html_parts:
            html = html.replace("<body>", '<body style="page-break-before: always;">')
        all_html_parts.append(html)
        print(f"   ✓ {os.path.basename(xlsx_path)} → {data['date_str']}（{len(data['tasks'])} 項）")

    # Merge into one document
    combined_html = all_html_parts[0]
    # For subsequent files, extract body content and inject
    for part in all_html_parts[1:]:
        # Extract everything inside <body>...</body>
        m = re.search(r"<body[^>]*>(.*?)</body>", part, re.DOTALL)
        if m:
            body_content = m.group(1)
            combined_html = combined_html.replace("</body>", body_content + "\n</body>")

    doc = weasyprint.HTML(string=combined_html)
    doc.write_pdf(output_path)
    print(f"✅ 已輸出合併 PDF：{output_path}")
    return output_path


# ════════════════════════════════════════════════════════════════
# 功能三：插入圖片到 XLSX 指定欄位
# ════════════════════════════════════════════════════════════════


def insert_image_into_xlsx(
    xlsx_path: str,
    image_path: str,
    task_name: str = None,
    row: int = None,
    col: int = None,
    max_width: int = 120,
    max_height: int = 120
):
    """
    Insert an image into a specific cell of a daily log XLSX.

    Args:
        xlsx_path: Path to the target XLSX file
        image_path: Path to the image file (JPEG/PNG)
        task_name: Insert at the row matching this task description
        row: Explicit row number (overrides task_name if both given)
        col: Column number (default None = auto-detect "照片" column)
        max_width/max_height: Image display size in pixels
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from PIL import Image as PilImage

    wb = load_workbook(xlsx_path)
    ws = wb["時間軸"]

    # Auto-detect photo column: 以欄位名「照片」定位（支援網站動態輸出模板）
    photo_col = col
    if not photo_col:
        _hr, col_map = detect_header_row(ws)
        photo_col = col_map.get("photo_status")
        if photo_col is None:
            # Fallback: 8-col vs 7-col vs 6-col heuristic
            header_4 = str(ws.cell(row=3, column=4).value or "")
            header_5 = str(ws.cell(row=3, column=5).value or "")
            is_7col = "分類" in header_4
            is_8col = is_7col and ("項位" in header_5 or "分類" in header_5)
            photo_col = 8 if is_8col else (7 if is_7col else 6)

    # Determine target row
    target_row = row
    if target_row is None and task_name:
        for r in range(4, ws.max_row + 1):  # data starts at row 4
            desc = str(ws.cell(row=r, column=2).value or "").strip()
            if desc == task_name:
                target_row = r
                break
        if target_row is None:
            print(f"❌ 找不到事項「{task_name}」")
            wb.close()
            return None

    if target_row is None:
        print("❌ 請指定 --task 或 --row")
        wb.close()
        return None

    # Verify image file
    if not os.path.isfile(image_path):
        print(f"❌ 找不到圖片：{image_path}")
        wb.close()
        return None

    # Resize image for XLSX
    pil_img = PilImage.open(image_path)
    pil_img.thumbnail((max_width, max_height), PilImage.LANCZOS)

    # Save temp resized image
    ext = os.path.splitext(image_path)[1].lower()
    temp_img = image_path + ".tmp" + ext
    pil_img.save(temp_img)

    # Insert into XLSX
    img = XlImage(temp_img)
    img.width = max_width
    img.height = max_height
    cell_ref = f"{chr(64 + photo_col)}{target_row}"
    img.anchor = cell_ref
    ws.add_image(img)

    # Update photo_status
    current_status = str(ws.cell(row=target_row, column=photo_col).value or "")
    if "無照片" in current_status:
        ws.cell(row=target_row, column=photo_col).value = "有拍照"

    wb.save(xlsx_path)
    wb.close()

    # Cleanup temp
    try:
        os.remove(temp_img)
    except OSError:
        pass

    task_desc = ws.cell(row=target_row, column=2).value or ""
    print(f"✅ 已插入圖片到「{task_desc}」({cell_ref})：{os.path.basename(image_path)}")
    return xlsx_path


# ── 入口 ──────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="工作日誌 — 輸入事項工具")

    # ── 現有功能 ──
    parser.add_argument(
        "--export-xlsx", "-x",
        metavar="MARKDOWN_FILE",
        help="將工作日誌 Markdown 檔案轉換為 XLSX"
    )

    # ── 功能一：XLSX → PDF ──
    parser.add_argument(
        "--xlsx-to-pdf", "-p",
        metavar="XLSX_FILE",
        help="將工作日誌 XLSX 轉換為 PDF"
    )

    # ── 功能二：多 XLSX 合併為 PDF ──
    parser.add_argument(
        "--merge-xlsx-pdf", "-m",
        nargs="+",
        metavar=("XLSX_FILE1", "XLSX_FILE2"),
        help="合併多個工作日誌 XLSX 檔案為單一 PDF"
    )

    # ── 功能三：插入圖片到 XLSX ──
    parser.add_argument(
        "--insert-image", "-i",
        nargs=2,
        metavar=("XLSX_FILE", "IMAGE_PATH"),
        help="將圖片插入 XLSX 指定欄位。需配合 --task 或 --row 使用"
    )
    parser.add_argument(
        "--task", "-t",
        metavar="TASK_NAME",
        help="配合 --insert-image 使用，指定事項名稱（如「取閘紙」）"
    )
    parser.add_argument(
        "--row", "-r",
        type=int,
        help="配合 --insert-image 使用，指定行號（覆蓋 --task）"
    )
    parser.add_argument(
        "--col", "-c",
        type=int,
        default=None,
        help="配合 --insert-image 使用，指定插入列號（預設自動檢測「照片」欄）"
    )

    # ── 輸出位置（可選） ──
    parser.add_argument(
        "--output", "-o",
        metavar="OUTPUT_PATH",
        help="指定輸出檔案路徑（支援 PDF/XLSX 類別）"
    )

    args = parser.parse_args()

    # ── 路由 ──

    if args.xlsx_to_pdf:
        xlsx_path = args.xlsx_to_pdf
        if not os.path.isfile(xlsx_path):
            print(f"❌ 找不到檔案：{xlsx_path}")
            sys.exit(1)
        xlsx_to_pdf(xlsx_path, args.output)
        sys.exit(0)

    if args.merge_xlsx_pdf:
        paths = args.merge_xlsx_pdf
        missing = [p for p in paths if not os.path.isfile(p)]
        if missing:
            print(f"❌ 找不到檔案：{', '.join(missing)}")
            sys.exit(1)
        merge_xlsx_pdf(paths, args.output)
        sys.exit(0)

    if args.insert_image:
        xlsx_path, image_path = args.insert_image
        if not os.path.isfile(xlsx_path):
            print(f"❌ 找不到 XLSX：{xlsx_path}")
            sys.exit(1)
        if not os.path.isfile(image_path):
            print(f"❌ 找不到圖片：{image_path}")
            sys.exit(1)
        insert_image_into_xlsx(
            xlsx_path, image_path,
            task_name=args.task,
            row=args.row,
            col=args.col
        )
        sys.exit(0)

    if args.export_xlsx:
        md_path = args.export_xlsx
        if not os.path.isfile(md_path):
            print(f"❌ 找不到檔案：{md_path}")
            sys.exit(1)

        print(f"📖 讀取：{md_path}")
        data = parse_markdown_log(md_path)
        print(f"   ✓ 日期：{data['date_str']}（{data['weekday']}）")
        print(f"   ✓ 人員：{data['person']}")
        print(f"   ✓ 事項：{len(data['tasks'])} 項")

        xlsx_path = os.path.splitext(md_path)[0] + ".xlsx"
        generate_xlsx(data, xlsx_path)
        print(f"✅ 已輸出：{xlsx_path}")
        sys.exit(0)

    app = DailyLogApp()
    app.run()
